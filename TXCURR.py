import pandas as pd 
import streamlit as st 
import os
import numpy as np
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import time
from pathlib import Path
#from streamlit_gsheets import GSheetsConnection
from datetime import datetime

st.set_page_config(
    page_title = 'MOCK TX CURR',
    page_icon =":bar_chart"
    )

#st.header('CODE UNDER MAINTENANCE, TRY AGAIN TOMORROW')
#st.stop()
cola,colb,colc = st.columns([1,3,1])
colb.subheader('PROGRAM GROWTH')

today = datetime.now()
todayd = today.strftime("%Y-%m-%d")# %H:%M")
wk = today.strftime("%V")
week = int(wk)-39
cola,colb = st.columns(2)
cola.write(f"**DATE TODAY:    {todayd}**")
colb.write(f"**CURRENT WEEK:    {week}**")

# HTML Table
html_table = """
<table style="width:100%">
  <tr>
    <th >EMR COLUMN</th>
    <th>RENAME TO:</th> 
    <th>EMR COLUMN</th>
    <th>RENAME TO:</th>
  </tr>
 <tr>
    <td><b>1. HIV Clinic No.</b></td>
    <td><b>ART</b></td>
    <td><b>2. ART START DATE</b></td>
     <td><b>AS</b></td>
 </tr>
 <tr>
    <td><b>3. TRANSFER OUT DATE</b></td>
    <td><b>TO</b></td>
    <td><b>4. Death Date</b></td>
     <td><b>DD</b></td>
 </tr>
  <tr>
    <td><b>5. LAST ENCOUNTER DATE</b></td>
    <td><b>LD</b></td>
    <td><b>6. FIRST ENCOUNTER DATE</b></td>
     <td><b>FE</b></td>
 </tr>
  <tr>
    <td><b>7. RETURN VISIT DATE</b></td>
    <td><b>RD</b></td>
    <td><b>8. RETURN VISIT DATE1</b></td>
     <td><b>RD1</b></td>
 </tr>
  <tr>
    <td><b>9.RETURN VISIT DATE2</b></td>
    <td><b>RD2</b></td>
    <td><b>10.RETURN VISIT DATE_Obs Date</b></td>
     <td><b>RDO</b></td>
 </tr>
  <tr>
    <td><b>11. TRANSFER IN OBS DATE</b></td>
    <td><b>TI</b></td>
    <td><b>12. HIV VIRAL LOAD DATE </b></td>
     <td><b>VD</b></td>
 </tr>
   <tr>
    <td><b>13. ARV REGIMEN DAYS DISPENSED</b></td>
    <td><b>ARVD</b></td>
    <td><b>14. ARV REGIMEN DAYS DISPENSED_obsDatetime</b></td>
     <td><b>ARVDO</b></td>
 </tr>
</table>
"""
# Display the HTML table using markdown in Streamlit
st.markdown(html_table, unsafe_allow_html=True)
html_table = """
<table style="width:100%">
  <tr>
    <th>AFTER, SAVE THIS EXTRACT AS an XLSX BEFORE YOU PROCEED</th>
  </tr>
</table>
"""

# Display the HTML table using markdown in Streamlit
#st.markdown(html_table, unsafe_allow_html=True)
st.markdown('**AFTER, SAVE THE EXTRACT AS an XLSX BEFORE YOU PROCEED**')

file = st.file_uploader("Upload your EMR extract here", type=['csv', 'xlsx'])
if 'submited' not in st.session_state:
    st.session_state.submited =False
ext = None
if file is not None:
    # Get the file name
    fileN = file.name
    ext = os.path.basename(fileN).split('.')[1]
df = None
if file is not None:
    if ext !='xlsx':
        st.write('Unsupported file format, first save the excel as xlsx and try again')
        st.stop()
    else:
        df = pd.read_excel(file)
        st.write('Excel accepted')
    if df is not None:
        columns = ['ART','AS', 'VD', 'RD','TO', 'TI', 'DD', 'FE','LD', 'RD1', 'RD2', 'RDO', 'ARVD', 'ARVDO']
        cols = df.columns.to_list()
        if not all(column in cols for column in columns):
            missing_columns = [column for column in columns if column not in cols]
            for column in missing_columns:
                st.markdown(f' **ERROR !!! {column} is not in the file uploaded**')
                st.markdown('**First rename all the columns as guided above**')
                st.stop()
        else:
              # Convert 'ART' column to string and create 'ART' column with numeric part to remove blanks
            df = df[['ART','AS', 'VD', 'RD','TO', 'TI', 'DD', 'FE','LD', 'RD1', 'RD2', 'RDO', 'ARVD', 'ARVDO']].copy()
            df['ART'] = df['ART'].astype(str)
            df['A'] = df['ART'].str.replace('[^0-9]', '', regex=True)
            df['A'] = pd.to_numeric(df['A'], errors= 'coerce')
            df = df[df['A']>0].copy()
            #df.dropna(subset='ART', inplace=True)
            
            df[['AS', 'RD', 'VD','TO','TI']] = df[['AS', 'RD', 'VD','TO','TI']].astype(str)
            if df['TI'].str.contains('YES').any():
                st.write("You may be using the Transfer in column instead of the Transfer_in Obs date column")
                st.stop()
            
            df['AS'] = df['AS'].astype(str)
            df['ARVD'] = df['ARVD'].astype(str)
            df['ARVDO'] = df['ARVDO'].astype(str)
            df['RD'] = df['RD'].astype(str)
            df['RD1'] = df['RD1'].astype(str)
            df['RD2'] = df['RD2'].astype(str)
            df['RDO'] = df['RDO'].astype(str)
            df['TI'] = df['TI'].astype(str)
            df['TO'] = df['TO'].astype(str)
            df['VD'] = df['VD'].astype(str)
            df['DD'] = df['DD'].astype(str)
            df['LD'] = df['LD'].astype(str)
            df['FE'] = df['FE'].astype(str)
            
            y = pd.DataFrame({'ART' :['2','3','4'], 'TI':['1-1-1',1,'1/1/1'], 'RD':['1-1-1',1,'1/1/1'],'DD':['1-1-1',1,'1/1/1'], 
                              'TO':['1-1-1',1,'1/1/1'], 'AS':['1-1-1',1,'1/1/1'], 'VD':['1-1-1',1,'1/1/1'],'RD1':['1-1-1',1,'1/1/1'],
                              'RD2':['1-1-1',1,'1/1/1'],'RDO':['1-1-1',1,'1/1/1'], 'ARVD':['1-1-1',1,'1/1/1'], 'ARVDO':['1-1-1',1,'1/1/1'],
                              'LD':['1-1-1',1,'1/1/1'],'FE':['1-1-1',1,'1/1/1']})  
            
            
            df['AS'] = df['AS'].astype(str)
            df['ARVDO'] = df['ARVDO'].astype(str)
            df['RD'] = df['RD'].astype(str)
            df['RD1'] = df['RD1'].astype(str)
            df['RD2'] = df['RD2'].astype(str)
            df['RDO'] = df['RDO'].astype(str)
            df['TI'] = df['TI'].astype(str)
            df['TO'] = df['TO'].astype(str)
            df['VD'] = df['VD'].astype(str)
            df['DD'] = df['DD'].astype(str)
            df['LD'] = df['LD'].astype(str)
            df['FE'] = df['FE'].astype(str)

            df['AS'] = df['AS'].str.replace('00:00:00', '', regex=True)
            df['ARVDO'] = df['ARVDO'].str.replace('00:00:00', '', regex=True)
            df['RD'] = df['RD'].str.replace('00:00:00', '', regex=True)
            df['RD1'] = df['RD1'].str.replace('00:00:00', '', regex=True)
            df['RD2'] = df['RD2'].str.replace('00:00:00', '', regex=True)
            df['RDO'] = df['RDO'].str.replace('00:00:00', '', regex=True)
            df['TI'] = df['TI'].str.replace('00:00:00', '', regex=True)
            df['TO'] = df['TO'].str.replace('00:00:00', '', regex=True)
            df['VD'] = df['VD'].str.replace('00:00:00', '', regex=True)
            df['DD'] = df['DD'].str.replace('00:00:00', '', regex=True)
            df['LD'] = df['LD'].str.replace('00:00:00', '', regex=True)
            df['FE'] = df['FE'].str.replace('00:00:00', '', regex=True)


            df = pd.concat([df,y])


            df['AS'] = df['AS'].astype(str) ###
            df['ARVDO'] = df['ARVDO'].astype(str)
            df['RD'] = df['RD'].astype(str) ###
            df['RD1'] = df['RD1'].astype(str)##
            df['RD2'] = df['RD2'].astype(str)##
            df['RDO'] = df['RDO'].astype(str)
            df['TI'] = df['TI'].astype(str) ##
            df['TO'] = df['TO'].astype(str) ##
            df['VD'] = df['VD'].astype(str) ###
            df['DD'] = df['DD'].astype(str) ####
            df['LD'] = df['LD'].astype(str)
            df['FE'] = df['FE'].astype(str)


            # SPLITTING ART START DATE
            A = df[df['AS'].str.contains('-')].copy()
            a = df[~df['AS'].str.contains('-')].copy()
            B = a[a['AS'].str.contains('/')].copy()
            C = a[~a['AS'].str.contains('/')].copy()

            A[['Ayear', 'Amonth', 'Aday']] = A['AS'].str.split('-', expand = True)
            B[['Ayear', 'Amonth', 'Aday']] = B['AS'].str.split('/', expand = True)
            try:            
                C['AS'] = pd.to_numeric(C['AS'], errors='coerce')
                C['AS'] = pd.to_datetime(C['AS'], origin='1899-12-30', unit='D', errors='ignore')
                C['AS'] =  C['AS'].astype(str)
                C[['Ayear', 'Amonth', 'Aday']] = C['AS'].str.split('-', expand = True)
            except:
                pass
            df = pd.concat([A,B,C])

             # SPLITTING DEATH DATE
            A = df[df['DD'].str.contains('-')].copy()
            a = df[~df['DD'].str.contains('-')].copy()
            B = a[a['DD'].str.contains('/')].copy()
            C = a[~a['DD'].str.contains('/')].copy()

            A[['Dyear', 'Dmonth', 'Dday']] = A['DD'].str.split('-', expand = True)
            B[['Dyear', 'Dmonth', 'Dday']] = B['DD'].str.split('/', expand = True)
            try:            
                C['DD'] = pd.to_numeric(C['DD'], errors='coerce')
                C['DD'] = pd.to_datetime(C['DD'], origin='1899-12-30', unit='D', errors='coerce')
                C['DD'] =  C['DD'].astype(str)
                C[['Dyear', 'Dmonth', 'Dday']] = C['DD'].str.split('-', expand = True)
            except:
                pass
            df = pd.concat([A,B,C])
          
            # SORTING THE RETURN VISIT DATE
            A = df[df['RD'].str.contains('-')].copy()
            a = df[~df['RD'].str.contains('-')].copy()
            B = a[a['RD'].str.contains('/')].copy()
            C = a[~a['RD'].str.contains('/')].copy()
      
            A[['Ryear', 'Rmonth', 'Rday']] = A['RD'].str.split('-', expand = True)
            B[['Ryear', 'Rmonth', 'Rday']] = B['RD'].str.split('/', expand = True)
            try:
                C['RD'] = pd.to_numeric(C['RD'], errors='coerce')
                C['RD'] = pd.to_datetime(C['RD'], origin='1899-12-30', unit='D', errors='coerce')
                C['RD'] =  C['RD'].astype(str)
                C[['Ryear', 'Rmonth', 'Rday']] = C['RD'].str.split('-', expand = True)
            except:
                pass
            df = pd.concat([A,B,C]) 
          
            #SORTING THE VD DATE
            A = df[df['VD'].str.contains('-')].copy()
            a = df[~df['VD'].str.contains('-')].copy()
            B = a[a['VD'].str.contains('/')].copy()
            C = a[~a['VD'].str.contains('/')].copy()

            A[['Vyear', 'Vmonth', 'Vday']] = A['VD'].str.split('-', expand = True)
            B[['Vyear', 'Vmonth', 'Vday']] = B['VD'].str.split('/', expand = True)
            try:
                C['VD'] = pd.to_numeric(C['VD'], errors='coerce')
                C['VD'] = pd.to_datetime(C['VD'], origin='1899-12-30', unit='D', errors='coerce')
                C['VD'] =  C['VD'].astype(str)
                C[['Vyear', 'Vmonth', 'Vday']] = C['VD'].str.split('-', expand = True)
            except:
                pass
            df = pd.concat([A,B,C])

            #SORTING THE TO DATE
            A = df[df['TO'].str.contains('-')].copy()
            a = df[~df['TO'].str.contains('-')].copy()
            B = a[a['TO'].str.contains('/')].copy()
            C = a[~a['TO'].str.contains('/')].copy()

            A[['Tyear', 'Tmonth', 'Tday']] = A['TO'].str.split('-', expand = True)
            B[['Tyear', 'Tmonth', 'Tday']] = B['TO'].str.split('/', expand = True)
            try:            
                C['TO'] = pd.to_numeric(C['TO'], errors='coerce')
                C['TO'] = pd.to_datetime(C['TO'], origin='1899-12-30', unit='D', errors='coerce')
                C['TO'] =  C['TO'].astype(str)
                C[['Tyear', 'Tmonth', 'Tday']] = C['TO'].str.split('-', expand = True)
            except:
                pass
            df = pd.concat([A,B,C])
        

           #SORTING THE TI DATE
            A = df[df['TI'].str.contains('-')].copy()
            a = df[~df['TI'].str.contains('-')].copy()
            B = a[a['TI'].str.contains('/')].copy()
            C = a[~a['TI'].str.contains('/')].copy()

            A[['Tiyear', 'Timonth', 'Tiday']] = A['TI'].str.split('-', expand = True)
            B[['Tiyear', 'Timonth', 'Tiday']] = B['TI'].str.split('/', expand = True)
            try:            
                C['TI'] = pd.to_numeric(C['TI'], errors='coerce')
                C['TI'] = pd.to_datetime(C['TI'], origin='1899-12-30', unit='D', errors='coerce')
                C['TI'] =  C['TI'].astype(str)
                C[['Tiyear', 'Timonth', 'Tiday']] = C['TI'].str.split('-', expand = True)
            except:
                pass
            df = pd.concat([A,B,C])

            # SORTING THE RETURN VISIT DATE1
            A = df[df['RD1'].str.contains('-')].copy()
            a = df[~df['RD1'].str.contains('-')].copy()
            B = a[a['RD1'].str.contains('/')].copy()
            C = a[~a['RD1'].str.contains('/')].copy()
      
            A[['R1year', 'R1month', 'R1day']] = A['RD1'].str.split('-', expand = True)
            B[['R1year', 'R1month', 'R1day']] = B['RD1'].str.split('/', expand = True)
            try:
                C['RD1'] = pd.to_numeric(C['RD1'], errors='coerce')
                C['RD1'] = pd.to_datetime(C['RD1'], origin='1899-12-30', unit='D', errors='coerce')
                C['RD1'] =  C['RD1'].astype(str)
                C[['R1year', 'R1month', 'R1day']] = C['RD1'].str.split('-', expand = True)
            except:
                pass
            df = pd.concat([A,B,C]) 
          
            # SORTING THE RETURN VISIT DATE2
            A = df[df['RD2'].str.contains('-')].copy()
            a = df[~df['RD2'].str.contains('-')].copy()
            B = a[a['RD2'].str.contains('/')].copy()
            C = a[~a['RD2'].str.contains('/')].copy()
      
            A[['R2year', 'R2month', 'R2day']] = A['RD2'].str.split('-', expand = True)
            B[['R2year', 'R2month', 'R2day']] = B['RD2'].str.split('/', expand = True)
            try:
                C['RD2'] = pd.to_numeric(C['RD2'], errors='coerce')
                C['RD2'] = pd.to_datetime(C['RD2'], origin='1899-12-30', unit='D', errors='coerce')
                C['RD2'] =  C['RD2'].astype(str)
                C[['R2year', 'R2month', 'R2day']] = C['RD2'].str.split('-', expand = True)
            except:
                pass
            df = pd.concat([A,B,C])
        
            # SORTING THE RETURN VISIT OBS DATE
            A = df[df['RDO'].str.contains('-')].copy()
            a = df[~df['RDO'].str.contains('-')].copy()
            B = a[a['RDO'].str.contains('/')].copy()
            C = a[~a['RDO'].str.contains('/')].copy()
      
            A[['ROyear', 'ROmonth', 'ROday']] = A['RDO'].str.split('-', expand = True)
            B[['ROyear', 'ROmonth', 'ROday']] = B['RDO'].str.split('/', expand = True)
            try:
                C['RDO'] = pd.to_numeric(C['RDO'], errors='coerce')
                C['RDO'] = pd.to_datetime(C['RDO'], origin='1899-12-30', unit='D', errors='coerce')
                C['RDO'] =  C['RDO'].astype(str)
                C[['ROyear', 'ROmonth', 'ROday']] = C['RDO'].str.split('-', expand = True)
            except:
                pass
            df = pd.concat([A,B,C])

            # SORTING THE LAST ENCOUNTER DATES
            A = df[df['LD'].str.contains('-')].copy()
            a = df[~df['LD'].str.contains('-')].copy()
            B = a[a['LD'].str.contains('/')].copy()
            C = a[~a['LD'].str.contains('/')].copy()
      
            A[['Lyear', 'Lmonth', 'Lday']] = A['LD'].str.split('-', expand = True)
            B[['Lyear', 'Lmonth', 'Lday']] = B['LD'].str.split('/', expand = True)
            try:
                C['LD'] = pd.to_numeric(C['LD'], errors='coerce')
                C['LD'] = pd.to_datetime(C['LD'], origin='1899-12-30', unit='D', errors='coerce')
                C['LD'] =  C['LD'].astype(str)
                C[['Lyear', 'Lmonth', 'Lday']] = C['LD'].str.split('-', expand = True)
            except:
                pass
            df = pd.concat([A,B,C])
          
            # SORTING THE ARV DISPENSED DATES
            A = df[df['ARVDO'].str.contains('-')].copy()
            a = df[~df['ARVDO'].str.contains('-')].copy()
            B = a[a['ARVDO'].str.contains('/')].copy()
            C = a[~a['ARVDO'].str.contains('/')].copy()
      
            A[['Aryear', 'Armonth', 'Arday']] = A['ARVDO'].str.split('-', expand = True)
            B[['Aryear', 'Armonth', 'Arday']] = B['ARVDO'].str.split('/', expand = True)
            try:
                C['ARVDO'] = pd.to_numeric(C['ARVDO'], errors='coerce')
                C['ARVDO'] = pd.to_datetime(C['ARVDO'], origin='1899-12-30', unit='D', errors='coerce')
                C['ARVDO'] =  C['ARVDO'].astype(str)
                C[['Aryear', 'Armonth', 'Arday']] = C['ARVDO'].str.split('-', expand = True)
            except:
                pass
            df = pd.concat([A,B,C])

            # SORTING THE FIRST ENCOUNTER
            A = df[df['FE'].str.contains('-')].copy()
            a = df[~df['FE'].str.contains('-')].copy()
            B = a[a['FE'].str.contains('/')].copy()
            C = a[~a['FE'].str.contains('/')].copy()
      
            A[['Fyear', 'Fmonth', 'Fday']] = A['FE'].str.split('-', expand = True)
            B[['Fyear', 'Fmonth', 'Fday']] = B['FE'].str.split('/', expand = True)
            try:
                C['FE'] = pd.to_numeric(C['FE'], errors='coerce')
                C['FE'] = pd.to_datetime(C['FE'], origin='1899-12-30', unit='D', errors='coerce')
                C['FE'] =  C['FE'].astype(str)
                C[['Fyear', 'Fmonth', 'Fday']] = C['FE'].str.split('-', expand = True)
            except:
                pass
            df = pd.concat([A,B,C])

               #BRINGING BACK THE / IN DATES
            df['AS'] = df['AS'].astype(str)
            df['ARVDO'] = df['ARVDO'].astype(str)
            df['RD'] = df['RD'].astype(str)
            df['RD1'] = df['RD1'].astype(str)
            df['RD2'] = df['RD2'].astype(str)
            df['RDO'] = df['RDO'].astype(str)
            df['TI'] = df['TI'].astype(str)
            df['TO'] = df['TO'].astype(str)
            df['VD'] = df['VD'].astype(str)
            df['DD'] = df['DD'].astype(str)
            df['LD'] = df['LD'].astype(str)
            df['FE'] = df['FE'].astype(str)

#             #Clearing NaT from te dates
            df['AS'] = df['AS'].str.replace('NaT', '',regex=True)
            df['ARVDO'] = df['ARVDO'].str.replace('NaT', '',regex=True)
            df['RD'] = df['RD'].str.replace('NaT', '',regex=True)
            df['RD1'] = df['RD1'].str.replace('NaT', '',regex=True)
            df['RD2'] = df['RD2'].str.replace('NaT', '',regex=True)
            df['RDO'] = df['RDO'].str.replace('NaT', '',regex=True)
            df['TI'] = df['TI'].str.replace('NaT', '',regex=True)
            df['TO'] = df['TO'].str.replace('NaT', '',regex=True)
            df['VD'] = df['VD'].str.replace('NaT', '',regex=True)
            df['DD'] = df['DD'].str.replace('NaT', '',regex=True)
            df['LD'] = df['LD'].str.replace('NaT', '',regex=True)
            df['FE'] = df['FE'].str.replace('NaT', '',regex=True)

                        #SORTING THE VIRAL LOAD YEARS
          
            df[['Vyear', 'Vmonth', 'Vday']] =df[['Vyear', 'Vmonth', 'Vday']].apply(pd.to_numeric, errors = 'coerce') 
            df['Vyear'] = df['Vyear'].fillna(994)
            a = df[df['Vyear']>31].copy()
            b = df[df['Vyear']<32].copy()
            b = b.rename(columns={'Vyear': 'Vday2', 'Vday': 'Vyear'})
            b = b.rename(columns={'Vday2': 'Vday'})
            df = pd.concat([a,b])
            dfa = df.shape[0]


             #SORTING THE TI YEARS
            df[['Tiyear', 'Tiday']] =df[['Tiyear','Tiday']].apply(pd.to_numeric, errors = 'coerce')
            df['Tiyear'] = df['Tiyear'].fillna(994)
            a = df[df['Tiyear']>31].copy()
            b = df[df['Tiyear']<32].copy()
            b = b.rename(columns={'Tiyear': 'Tiday2', 'Tiday': 'Tiyear'})
            b = b.rename(columns={'Tiday2': 'Tiday'})
            df = pd.concat([a,b])
            dfb = df.shape[0]

            # #SORTING THE RETURN VISIT DATE YEARS
            df[['Rday', 'Ryear']] = df[['Rday', 'Ryear']].apply(pd.to_numeric, errors='coerce')
            
            df['Ryear'] = df['Ryear'].fillna(994)
            a = df[df['Ryear']>31].copy()
            b = df[df['Ryear']<32].copy()
            b = b.rename(columns={'Ryear': 'Rday2', 'Rday': 'Ryear'})
            b = b.rename(columns={'Rday2': 'Rday'})

            df = pd.concat([a,b])
            dfc = df.shape[0]
            
                #SORTING THE TRANSFER OUT DATE YEAR
            df[['Tday', 'Tyear']] = df[['Tday', 'Tyear']].apply(pd.to_numeric, errors='coerce')
            df['Tyear'] = df['Tyear'].fillna(994)
            a = df[df['Tyear']>31].copy()
            b = df[df['Tyear']<32].copy()
            b = b.rename(columns={'Tyear': 'Tday2', 'Tday': 'Tyear'})
            b = b.rename(columns={'Tday2': 'Tday'})
            df = pd.concat([a,b])

            
               #SORTING THE ART START YEARS
            df[['Ayear', 'Amonth', 'Aday']] =df[['Ayear', 'Amonth', 'Aday']].apply(pd.to_numeric, errors = 'coerce')
            df['Ayear'] = df['Ayear'].fillna(994)
            a = df[df['Ayear']>31].copy()
            b = df[df['Ayear']<32].copy()
            b = b.rename(columns={'Ayear': 'Aday2', 'Aday': 'Ayear'})
            b = b.rename(columns={'Aday2': 'Aday'})
            df = pd.concat([a,b])
            dfe = df.shape[0]

              #SORTING THE ART START YEARS
            df[['Dyear', 'Dmonth', 'Dday']] =df[['Dyear', 'Dmonth', 'Dday']].apply(pd.to_numeric, errors = 'coerce')
            df['Dyear'] = df['Dyear'].fillna(994)
            a = df[df['Dyear']>31].copy()
            b = df[df['Dyear']<32].copy()
            b = b.rename(columns={'Dyear': 'Dday2', 'Dday': 'Dyear'})
            b = b.rename(columns={'Dday2': 'Dday'})
            df = pd.concat([a,b])
            dfe = df.shape[0]

            # #SORTING THE RETURN VISIT DATE1
            df[['R1day', 'R1year']] = df[['R1day', 'R1year']].apply(pd.to_numeric, errors='coerce')
            
            df['R1year'] = df['R1year'].fillna(994)
            a = df[df['R1year']>31].copy()
            b = df[df['R1year']<32].copy()
            b = b.rename(columns={'R1year': 'R1day2', 'R1day': 'R1year'})
            b = b.rename(columns={'R1day2': 'R1day'})

            df = pd.concat([a,b])
            dfc = df.shape[0]

            # #SORTING THE RETURN VISIT DATE2
            df[['R2day', 'R2year']] = df[['R2day', 'R2year']].apply(pd.to_numeric, errors='coerce')
            
            df['R2year'] = df['R2year'].fillna(994)
            a = df[df['R2year']>31].copy()
            b = df[df['R2year']<32].copy()
            b = b.rename(columns={'R2year': 'R2day2', 'R2day': 'R2year'})
            b = b.rename(columns={'R2day2': 'R2day'})

            df = pd.concat([a,b])
            dfc = df.shape[0]

            # #SORTING THE RETURN VISIT OBS DATE
            df[['ROday', 'ROyear']] = df[['ROday', 'ROyear']].apply(pd.to_numeric, errors='coerce')
            
            df['ROyear'] = df['ROyear'].fillna(994)
            a = df[df['ROyear']>31].copy()
            b = df[df['ROyear']<32].copy()
            b = b.rename(columns={'ROyear': 'ROday2', 'ROday': 'ROyear'})
            b = b.rename(columns={'ROday2': 'ROday'})

            df = pd.concat([a,b])
            dfc = df.shape[0]

            # #SORTING THE LAST ENCOUNTER
            df[['Lday', 'Lyear']] = df[['Lday', 'Lyear']].apply(pd.to_numeric, errors='coerce')
            
            df['Lyear'] = df['Lyear'].fillna(994)
            a = df[df['Lyear']>31].copy()
            b = df[df['Lyear']<32].copy()
            b = b.rename(columns={'Lyear': 'Lday2', 'Lday': 'Lyear'})
            b = b.rename(columns={'Lday2': 'Lday'})

            df = pd.concat([a,b])
            dfc = df.shape[0]

            # #SORTING THE FIRST ENCOUNTER
            df[['Fday', 'Fyear']] = df[['Fday', 'Fyear']].apply(pd.to_numeric, errors='coerce')
            
            df['Fyear'] = df['Fyear'].fillna(994)
            a = df[df['Fyear']>31].copy()
            b = df[df['Fyear']<32].copy()
            b = b.rename(columns={'Fyear': 'Fday2', 'Fday': 'Fyear'})
            b = b.rename(columns={'Fday2': 'Fday'})

            df = pd.concat([a,b])
            dfc = df.shape[0]

            # #SORTING THE FIRST ENCOUNTER
            df[['Arday', 'Aryear']] = df[['Arday', 'Aryear']].apply(pd.to_numeric, errors='coerce')
            
            df['Aryear'] = df['Aryear'].fillna(994)
            a = df[df['Aryear']>31].copy()
            b = df[df['Aryear']<32].copy()
            b = b.rename(columns={'Aryear': 'Arday2', 'Arday': 'Aryear'})
            b = b.rename(columns={'Arday2': 'Arday'})
            df = pd.concat([a,b])
            dfc = df.shape [0]

            #CREATE WEEKS 
            df['Rdaya'] = df['Rday'].astype(str).str.split('.').str[0]
            df['Rmontha'] = df['Rmonth'].astype(str).str.split('.').str[0]
            df['Ryeara'] = df['Ryear'].astype(str).str.split('.').str[0]

            df['RETURN DATE'] = df['Rdaya'] + '/' + df['Rmontha'] + '/' + df['Ryeara']
            df['RETURN DATE'] = pd.to_datetime(df['RETURN DATE'], format='%d/%m/%Y', errors='coerce')
            #CREATING WEEEK FOR RETURN VISIT DATE
            df['RWEEK'] = df['RETURN DATE'].dt.strftime('%V')
            df['RWEEK'] = pd.to_numeric(df['RWEEK'], errors='coerce')
            #SURGE WEEK
            df['RWEEK'] = pd.to_numeric(df['RWEEK'], errors='coerce')
            df['RWEEK1'] = df['RWEEK']-39
            #COPY FOR ONE YEAR BEFORE GETTING POT CURR
            oneyear = df.copy()
           
            file2 = r'CURRS.csv'
            dfx = pd.read_csv(file2)

            #POTENTIAL TXCUR ALTER... 
            df[['Rmonth', 'Rday', 'Ryear']] = df[['Rmonth', 'Rday', 'Ryear']].apply(pd.to_numeric, errors='coerce')
            df25 = df[df['Ryear']>2024].copy()
            df24 = df[df['Ryear'] == 2024].copy()
            df24[['Rmonth', 'Rday']] = df24[['Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
            df24 = df24[((df24['Rmonth']>9) | ((df24['Rmonth']==9) & (df24['Rday']>2)))].copy()
            df = pd.concat([df25, df24]).copy()

            #REMOVE TO of the last reporting month
            df[ 'Tyear'] = pd.to_numeric(df['Tyear'], errors='coerce')
            dfto = df[df['Tyear']!=994].copy()
            dfnot = df[df['Tyear'] == 994].copy()
            dfto[['Ryear', 'Rmonth']] = dfto[['Ryear', 'Rmonth']].apply(pd.to_numeric, errors='coerce')
            dfto = dfto[((dfto['Ryear']!=2024) |((dfto['Ryear']==2024) & (dfto['Rmonth']!=9)))].copy()
            df = pd.concat([dfto,dfnot])

            #REMOVE TO of the dead reporting month
            df[ 'Dyear'] = pd.to_numeric(df['Dyear'], errors='coerce')
            dfdd = df[df['Dyear']!=994].copy()
            dfnot = df[df['Dyear'] == 994].copy()
            #THOSE WHO DIED BEFORE FIRST MONTH OF THE Q
            dfdd[['Dyear', 'Dmonth']] = dfdd[['Dyear', 'Dmonth']].apply(pd.to_numeric, errors='coerce')
            dfdd = dfdd[((dfdd['Dyear']>2024) |((dfdd['Dyear']==2024) & (dfdd['Dmonth']>9)))].copy()
            df = pd.concat([dfdd,dfnot])
            pot = df.shape[0]

            #TRANSFER OUTS
            
            #TRANSFER INS
            df[['Tiyear', 'Timonth']] = df[['Tiyear', 'Timonth']].apply(pd.to_numeric, errors='coerce')
            dfti = df[((df['Tiyear']==2024) & (df['Timonth']>9))].copy() #TI
            ti = dfti.shape[0]

            dfnot = df[((df['Tiyear']!=2024) | ((df['Tiyear']==2024) & (df['Timonth']<10)))].copy() #NO TI
            noti = dfnot.shape[0]

            #TX NEW THIS Q
            dfnot[['Ayear', 'Amonth']] = dfnot[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
            dfnew = dfnot[((dfnot['Ayear']==2024) & (dfnot['Amonth']>9))].copy() #TI
            txnew = dfnew.shape[0]

            dfold = dfnot[((dfnot['Ayear']!=2024) | ((dfnot['Ayear']==2024) & (dfnot['Amonth']<10)))].copy() #NO TI
            dfcheck = dfold.copy() #use this to determine unknown gain
            old = dfold.shape[0]
            ##RTT
            #RTT BY LAST ENCOUNTER to include only months in the reporting Q
            dfold['Lyear'] = pd.to_numeric(dfold['Lyear'], errors='coerce') 
            dfRTT = dfold[dfold['Lyear']==2024].copy() #ALTER
            dfRTT['Lmonth'] = pd.to_numeric(dfRTT['Lmonth'], errors='coerce') 
            dfRTT = dfRTT[dfRTT['Lmonth'].isin([10,11,12])].copy() #ALTER

            #BY FIRST ENCOUNTER, To remove those first encountered in the Q
            dfRTT['Fyear'] = pd.to_numeric(dfRTT['Fyear'], errors='coerce') 
            dfRTTa = dfRTT[dfRTT['Fyear']==2024].copy() #ALTER
            dfRTTb = dfRTT[dfRTT['Fyear']!=2024].copy() #ALTER
            #BY FIRST ENCOUNTER
            dfRTTa['Fmonth'] = pd.to_numeric(dfRTTa['Fmonth'], errors='coerce') 
            dfRTTa = dfRTTa[~dfRTTa['Fmonth'].isin([10,11,12])].copy() # ALTER
            dfRTT = pd.concat([dfRTTa, dfRTTb])

            # #BY ART START, To remove those that started ART in the Q
            # dfRTT['Ayear'] = pd.to_numeric(dfRTT['Ayear'], errors='coerce')
            # dfRTTa = dfRTT[dfRTT['Ayear']==2024].copy() #ALTER
            # dfRTTb = dfRTT[dfRTT['Ayear']!=2024].copy() #ALTER ##ALREADY REMOVED ABOVE
            # #BY ART START
            # dfRTTa['Amonth'] = pd.to_numeric(dfRTTa['Amonth'], errors='coerce') 
            # dfRTTa = dfRTTa[~dfRTTa['Amonth'].isin([10,11,12])].copy() # ALTER use <
            # dfRTT = pd.concat([dfRTTa, dfRTTb])
            # #BY TI DATE, To remove those that TI in the Q
            # dfRTT['Tiyear'] = pd.to_numeric(dfRTT['Tiyear'], errors='coerce')
            # dfRTTa = dfRTT[dfRTT['Tiyear']==2024].copy() #ALTER
            # dfRTTb = dfRTT[dfRTT['Tiyear']!=2024].copy() #ALTER
            # #BY TI 
            # dfRTTa['Timonth'] = pd.to_numeric(dfRTTa['Timonth'], errors='coerce') 
            # dfRTTa = dfRTTa[~dfRTTa['Timonth'].isin([7,8,9])].copy() # ALTER use <
            # dfRTT = pd.concat([dfRTTa, dfRTTb])

            #BY RD OBS DATE,  remove those that fall in the previous reporting Quarter
            dfRTT['ROyear'] = pd.to_numeric(dfRTT['ROyear'], errors='coerce')
            dfRTTa = dfRTT[dfRTT['ROyear']>2024].copy()
            dfRTTb = dfRTT[dfRTT['ROyear']==2024].copy() 
            dfRTTb[['ROmonth', 'ROday']] = dfRTTb[['ROmonth', 'ROday']].apply(pd.to_numeric, errors='coerce')
            dfRTTb = dfRTTb[((dfRTTb['ROmonth']>9) | ((dfRTTb['ROmonth']==9) & (dfRTTb['ROday']>2)))].copy()
            dfRTT = pd.concat([dfRTTa, dfRTTb])

            #BY RDDATE1,  take those that fall in the previous reporting Quarter
            dfRTT['R1year'] = pd.to_numeric(dfRTT['R1year'], errors='coerce') 
            dfRTTa = dfRTT[dfRTT['R1year']<2024].copy()
            dfRTTb = dfRTT[dfRTT['R1year']==2024].copy()
            dfRTTb[['R1month', 'R1day']] = dfRTTb[['R1month', 'R1day']].apply(pd.to_numeric, errors='coerce')
            dfRTTb = dfRTTb[((dfRTTb['R1month']>6) | ((dfRTTb['R1month']==6) & (dfRTTb['R1day']<3)))].copy()
            dfRTT = pd.concat([dfRTTa, dfRTTb])

            #BY RD DATE2,  take those that fall in the previous reporting Quarter
            dfRTT['R2year'] = pd.to_numeric(dfRTT['R2year'], errors='coerce')
            dfRTTa = dfRTT[dfRTT['R2year']<2024].copy()
            dfRTTb = dfRTT[dfRTT['R2year']==2024].copy()
            dfRTTb[['R2month', 'R2day']] = dfRTTb[['R2month', 'R2day']].apply(pd.to_numeric, errors='coerce')
            dfRTTb = dfRTTb[((dfRTTb['R2month']>6) | ((dfRTTb['R2month']==6) & (dfRTTb['R2day']<3)))].copy()
            dfRTT = pd.concat([dfRTTa, dfRTTb])

            #BY ARV DISPENSED, to take those that got ART in the Q
            dfRTT['Aryear'] = pd.to_numeric(dfRTT['Aryear'], errors='coerce') 
            dfRTT = dfRTT[dfRTT['Aryear']==2024].copy() 
            dfRTT['Armonth'] = pd.to_numeric(dfRTT['Armonth'], errors='coerce') 
            dfRTT = dfRTT[dfRTT['Armonth'].isin([10,11,12])].copy()
            rtt = dfRTT.shape[0]
            #check

#######LOSSES. START FROM POTENTIAL CURR
           #TRANSFER OUTS
            df['Tyear'] = pd.to_numeric(df['Tyear'], errors='coerce')
            dfto = df[df['Ryear']==994].copy()
            dfnot = df[df['Ryear']!=994].copy()
            wk = int(wk)

            #FALSE TO OUTS BASED ON CURRENT WEEK
            dfto[['Ryear', 'RWEEK']] =  dfto[['Ryear', 'RWEEK']].apply(pd.to_numeric, errors='coerce')
            dfw = dfto[((dfto['Ryear']>2024) | ((dfto['Ryear']==2024) & (dfto['RWEEK']>=wk)))].copy() #FALSE
            false = dfw.shape[0]
            dft = dfto[((dfto['Ryear']<2024) | ((dfto['Ryear']==2024) & (dfto['RWEEK']<wk)))].copy()  ##TRUE
            true = dft.shape[0]
            #add the false back to txcur
            df = pd.concat([dfnot,dfw]) #WILL USE THIS FOR ACTIVE LATER

            #THOSE THAT HAVE DIED SO FAR
            df[ 'Dyear'] = pd.to_numeric(df['Dyear'], errors='coerce')
            dd = df[df['Dyear']!=994].copy() #DIED
            dead = dd.shape[0]

            #THIS CURR WILL HAVE NO DEAD AND TRUE TO
            df = df[df['Dyear'] == 994].copy() #LIVING, NO DEATH DATE

            #REMOVNG CURRENT LOST
            #USE CALENDAR WEEK FOR THIS Q, SWITCH TO SURGE WEEK NEXT Q
            #lost 2 weeks
            wk = int(wk)
            wk2 = wk-1
            wk3 = wk-2
            wk4 = wk-3
            df['Ryear'] = pd.to_numeric(df['Ryear'], errors='coerce')
            df24 = df[df['Ryear'] ==2024].copy()
            df25 = df[df['Ryear']>2024].copy()
            
            df24['RWEEK'] = pd.to_numeric(df24['RWEEK'], errors='coerce')
            dfactive24 =df24[df24['RWEEK']>=wk2] #still active within 2 weeks
            
            #LOST IN TWO WEEKS... REAL MISSED APPOINTMENT FOR THIS
            df2wks =df24[df24['RWEEK']<wk2].copy()
            two = df2wks.shape

            df3wks = df24[df24['RWEEK']<wk3]
            three = df3wks.shape

            df4wks =df24[df24['RWEEK1']<wk4]
            four = df4wks.shape

            dfactive = pd.concat([dfactive24, df25]) #COMBINE THOSE ACTIVE IN TWO WEEKS AND THOSE OF 2025
            active = dfactive.shape[0]

            #MMD AMONGST ACTIVE CLIENTS
            dfactive['ARVD'] = dfactive['ARVD'].fillna(20)
            dfactive['ARVD'] = pd.to_numeric(dfactive['ARVD'], errors='coerce')
            def mmd(a):
                if a<90:
                    return '3 MTHS'
                elif a< 180:
                    return '<6 MTHS'
                else:
                    return '6 MTHS+'
            dfactive = dfactive.copy() #avoid fragmentation
            dfactive['MULTI'] = dfactive['ARVD'].apply(mmd)
            dfactive['MULTI'] = dfactive['MULTI'].astype(str)
            df2mths =  dfactive[dfactive['MULTI']=='<3 MTHS'].copy()
            M2 = df2mths.shape[0]
            df3mths =  dfactive[dfactive['MULTI']=='<6 MTHS'].copy()
            M3 = df3mths.shape[0]
            df6mths =  dfactive[dfactive['MULTI']=='6 MTHS+'].copy()
            M6 = df6mths.shape[0]

            #VL SECTION
            #REMOVING SIX MONTHS TX NEW, to take those that got ART in the Q
            dfactive['Ayear'] = pd.to_numeric(dfactive['Ayear'], errors='coerce') 
            VLa = dfactive[dfactive['Ayear']<2024].copy()
            VLb = dfactive[dfactive['Ayear']==2024].copy()
            VLb = VLb[VLb['Amonth']<7].copy()
            VL = pd.concat([VLa,VLb])
            el = VL.shape[0]
            VL[['Vyear', 'Vmonth']] = VL[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
            WVL = VL[((VL['Vyear']>2023) | ((VL['Vyear']==2023) & (VL['Vmonth']>9)))].copy()
            NVL = VL[((VL['Vyear']<2023) | ((VL['Vyear']==2023) & (VL['Vmonth']<10)))].copy()
            nvl = NVL.shape[0]
            wvl = WVL.shape[0]

            #VL COV AMONG LOST CLIENTS
            df2wks['Ayear'] = pd.to_numeric(df2wks['Ayear'], errors='coerce')
            LVLa = df2wks[df2wks['Ayear']<2024].copy()
            LVLb = df2wks[df2wks['Ayear']==2024].copy()
            LVLb = LVLb[LVLb['Amonth']<7].copy()
            LVL = pd.concat([LVLa,LVLb])
            Lostelig = LVL.shape[0]
            LVL[['Vyear', 'Vmonth']] = LVL[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
            LWVL = LVL[((LVL['Vyear']>2023) | ((LVL['Vyear']==2023) & (LVL['Vmonth']>9)))].copy()
            LNVL = LVL[((LVL['Vyear']<2023) | ((LVL['Vyear']==2023) & (LVL['Vmonth']<10)))].copy()
            lnvl = LNVL.shape[0]
            lwvl = LWVL.shape[0]
            totalvl = pd.concat([LNVL,NVL])

            #EARLY RETENTION
            #ONE YEAR COHORT

            oneyear[['Ayear', 'Amonth']] = oneyear[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
            new = oneyear[((oneyear['Ayear']==2023) & (oneyear['Amonth'].isin([10,11,12])))].copy()
            newtotal = new.shape[0]

            new[['Tiyear']] = new[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
            tin = new[new['Tiyear']!=994].copy()
            #one =new.shape[0]
            newti = tin.shape[0]
            
            new['Dyear'] = pd.to_numeric(new['Dyear'], errors='coerce')
            newdead = new[new['Dyear']!=994].copy()

            deadnew = newdead.shape[0]
            new = new[new['Dyear']==994].copy() #AFTER REMOVING THE DEAD

            new['Tyear'] = pd.to_numeric(new['Tyear'], errors='coerce')
            
            newto = new[new['Tyear']!=994].copy()
            outnew = newto.shape[0]
            
            new = new[new['Tyear']==994].copy() #withou TO
            netnew = int(newtotal)- int(outnew)

            new['A'] = pd.to_numeric(new['A'], errors = 'coerce')
            dfactive['A'] = pd.to_numeric(dfactive['A'], errors = 'coerce')
            
            active = new[new['A'].isin(dfactive['ART'])].copy()
            lostn = new[~new['A'].isin(dfactive['A'])].copy()           

            newactive = active.shape[0]
            newlost = lostn.shape[0]
            #st.write(newlost)
            #VL SECTION AT ONE YEAR
            active[['Vyear', 'Vmonth']] = active[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
            WVLa = active[ ((active['Vyear']==2024) & (active['Vmonth']>9))].copy()
            NVLa = active[((active['Vyear']<2024) | ((active['Vyear']==2024) & (active['Vmonth']<10)))].copy()
            nvla = NVLa.shape[0]
            wvla = WVLa.shape[0]
                    
            #ret = newtotal - newlost
            if netnew == 0:
                rete = 0
            elif newactive == 0:
                rete = 0
            else:
                rete = round((newactive/netnew)*100)
                rete = f"{rete} %"


           #6 MONTH COHORT

            oneyear[['Ayear', 'Amonth']] = oneyear[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
            new6 = oneyear[((oneyear['Ayear']==2024) & (oneyear['Amonth'].isin([4,5,6])))].copy()
            newtotal6 = new6.shape[0]

            new6[['Tiyear']] = new6[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
            tin6 = new6[new6['Tiyear']!=994].copy()
            #one =new.shape[0]
            newti6 = tin6.shape[0]
            
            new6['Dyear'] = pd.to_numeric(new6['Dyear'], errors='coerce')
            newdead6 = new6[new6['Dyear']!=994].copy()

            deadnew6 = newdead6.shape[0]
            new6 = new6[new6['Dyear']==994].copy() #AFTER REMOVING THE DEAD

            new6['Tyear'] = pd.to_numeric(new6['Tyear'], errors='coerce')
            
            newto6 = new6[new6['Tyear']!=994].copy()
            outnew6 = newto6.shape[0]
            
            new6 = new6[new6['Tyear']==994].copy() #withou TO
            netnew6 = int(newtotal6)- int(outnew6)

            new6['A'] = pd.to_numeric(new6['A'], errors = 'coerce')
            dfactive['A'] = pd.to_numeric(dfactive['A'], errors = 'coerce')
            
            active6 = new6[new6['A'].isin(dfactive['ART'])].copy()
            lostn6 = new6[~new6['A'].isin(dfactive['A'])].copy()
            

            newactive6 = active6.shape[0]
            newlost6 = lostn6.shape[0]
            #st.write(newlost)
            #VL SECTION AT 6 MONTHS
            active6[['Vyear', 'Vmonth']] = active6[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
            WVLa6 = active[active['Vyear']==2024].copy()
            NVLa6 = active[active['Vyear']!=2024].copy()
            nvla6 = NVLa6.shape[0]
            wvla6 = WVLa6.shape[0]
            #ret = newtotal - newlost
            if netnew6 == 0:
                rete6 = 0
            elif newactive6 == 0:
                rete6 = 0
            else:
                rete6 = round((newactive6/netnew6)*100)
                rete6 = f"{rete6} %"

          #3 MONTH COHORT

            oneyear[['Ayear', 'Amonth']] = oneyear[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
            new3 = oneyear[((oneyear['Ayear']==2024) & (oneyear['Amonth'].isin([7,8,9])))].copy()
            newtotal3 = new3.shape[0]

            new3[['Tiyear']] = new3[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
            tin3 = new3[new3['Tiyear']!=994].copy()
            #one =new.shape[0]
            newti3 = tin3.shape[0]
            
            new3['Dyear'] = pd.to_numeric(new3['Dyear'], errors='coerce')
            newdead3 = new3[new3['Dyear']!=994].copy()

            deadnew3 = newdead3.shape[0]
            new3 = new3[new3['Dyear']==994].copy() #AFTER REMOVING THE DEAD

            new3['Tyear'] = pd.to_numeric(new3['Tyear'], errors='coerce')
            
            newto3 = new3[new3['Tyear']!=994].copy()
            outnew3 = newto3.shape[0]
            
            new3 = new3[new3['Tyear']==994].copy() #withou TO
            netnew3 = int(newtotal3)- int(outnew3)

            new3['A'] = pd.to_numeric(new3['A'], errors = 'coerce')
            dfactive['A'] = pd.to_numeric(dfactive['A'], errors = 'coerce')
            
            active3 = new3[new3['A'].isin(dfactive['ART'])].copy()
            lostn3 = new3[~new3['A'].isin(dfactive['A'])].copy()
            

            newactive3 = active3.shape[0]
            newlost3 = lostn3.shape[0]
            #st.write(newlost)
                    
            #ret = newtotal - newlost
            if netnew3 == 0:
                rete3 = 0
            elif newactive3 == 0:
                rete3 = 0
            else:
                rete3 = round((newactive3/netnew3)*100)
                rete3 = f"{rete3} %"

          #<3 MONTH COHORT

            oneyear[['Ayear', 'Amonth']] = oneyear[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
            new1 = oneyear[((oneyear['Ayear']==2024) & (oneyear['Amonth'].isin([10,11,12])))].copy()
            newtotal1 = new1.shape[0]

            new1[['Tiyear']] = new1[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
            tin1 = new1[new1['Tiyear']!=994].copy()
            #one =new.shape[0]
            newti1 = tin1.shape[0]
            
            new1['Dyear'] = pd.to_numeric(new1['Dyear'], errors='coerce')
            newdead1 = new1[new1['Dyear']!=994].copy()

            deadnew1 = newdead1.shape[0]
            new1 = new1[new1['Dyear']==994].copy() #AFTER REMOVING THE DEAD

            new1['Tyear'] = pd.to_numeric(new1['Tyear'], errors='coerce')
            
            newto1 = new1[new1['Tyear']!=994].copy()
            outnew1 = newto1.shape[0]
            
            new1 = new1[new1['Tyear']==994].copy() #withou TO
            netnew1 = int(newtotal1)- int(outnew1)

            new1['A'] = pd.to_numeric(new1['A'], errors = 'coerce')
            dfactive['A'] = pd.to_numeric(dfactive['A'], errors = 'coerce')
            
            active1 = new1[new1['A'].isin(dfactive['ART'])].copy()
            lostn1 = new1[~new1['A'].isin(dfactive['A'])].copy()
            

            newactive1 = active1.shape[0]
            newlost1 = lostn1.shape[0]
            #st.write(newlost)
                    
            #ret = newtotal - newlost
            if netnew1 == 0:
                rete1 = 0
            elif newactive1 == 0:
                rete1 = 0
            else:
                rete1 = round((newactive1/netnew1)*100)
                rete1 = f"{rete1} %"















#             dfy[['Rmonth', 'Rday']] = dfy[['Rmonth', 'Rday']].apply(pd.to_numeric, errors = 'coerce')
#             dfy = dfy[((dfy['Rmonth']>6) | ((dfy['Rmonth']==6) & (dfy['Rday'] >2)))].copy()
#             df = pd.concat([dfw,dfy])
#             potential = df.shape[0]
#             dpot = df.copy()
            
#             df['Dyear'] = pd.to_numeric(df['Dyear'], errors='coerce')
#             dead = df[df['Dyear']!=994].copy()

#             df = df[df['Dyear']==994].copy()
            
#             df['Rday1'] = df['Rday'].astype(str).str.split('.').str[0]
#             df['Rmonth1'] = df['Rmonth'].astype(str).str.split('.').str[0]
#             df['Ryear1'] = df['Ryear'].astype(str).str.split('.').str[0]

#             df['Vday1'] = df['Vday'].astype(str).str.split('.').str[0]
#             df['Vmonth1'] = df['Vmonth'].astype(str).str.split('.').str[0]
#             df['Vyear1'] = df['Vyear'].astype(str).str.split('.').str[0]

#             #df['Tiday'] = df['Tiday'].astype(str).str.split('.').str[0]
#             #df['Timonth'] = df['Timonth'].astype(str).str.split('.').str[0]
#             #df['Tiyear'] = df['Tiyear'].astype(str).str.split('.').str[0]

#             df['Aday1'] = df['Aday'].astype(str).str.split('.').str[0]
#             df['Amonth1'] = df['Amonth'].astype(str).str.split('.').str[0]
#             df['Ayear1'] = df['Ayear'].astype(str).str.split('.').str[0]
            
#             df['Tday1'] = df['Tday'].astype(str).str.split('.').str[0]
#             df['Tmonth1'] = df['Tmonth'].astype(str).str.split('.').str[0]
#             df['Tyear1'] = df['Tyear'].astype(str).str.split('.').str[0]

#             df['ART START DATE'] = df['Aday1'] + '/' + df['Amonth1'] + '/' + df['Ayear1']
#             df['RETURN DATE'] = df['Rday1'] + '/' + df['Rmonth1'] + '/' + df['Ryear1']
#             df['VL DATE'] = df['Vday1'] + '/' + df['Vmonth1'] + '/' + df['Vyear1']
#             df['T OUT DATE'] = df['Tday1'] + '/' + df['Tmonth1'] + '/' + df['Tyear1']
#             #df['T IN DATE'] = df['Rday1'] + '/' + df['Rmonth1'] + '/' + df['Ryear1']

#             df['RETURN DATE'] = pd.to_datetime(df['RETURN DATE'], format='%d/%m/%Y', errors='coerce')
#             df['VL DATE'] = pd.to_datetime(df['VL DATE'], format='%d/%m/%Y', errors='coerce')
#             df['T OUT DATE'] = pd.to_datetime(df['T OUT DATE'], format='%d/%m/%Y', errors='coerce')
#             df['ART START DATE'] = pd.to_datetime(df['ART START DATE'], format='%d/%m/%Y', errors='coerce')

#             df['RETURN DATE'] = df['RETURN DATE'].dt.strftime('%d/%m/%Y')
#             df['VL DATE'] = df['VL DATE'].dt.strftime('%d/%m/%Y')
#             df['T OUT DATE'] = df['T OUT DATE'].dt.strftime('%d/%m/%Y')
#             df['ART START DATE'] = df['ART START DATE'].dt.strftime('%d/%m/%Y')
            
#             df = df.rename(columns={'A': 'ART NO'})#, 'AS': 'ART START DATE', 'RD': 'RETURN DATE', 'VD': 'VL DATE', 'TO': 'T OUT DATE'})
            
#             df[['Tyear', 'Ryear', 'Rmonth', 'Rday', 'Vyear', 'Vmonth', 'Ayear']] = df[['Tyear', 'Ryear', 'Rmonth', 'Rday', 'Vyear', 'Vmonth', 'Ayear']].apply(pd.to_numeric, errors='coerce')
            
            
#             TXML = df[df['Ryear']==2024].copy()
#             TXML[['Rmonth', 'Rday']] = TXML[['Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
#             TXML = TXML[((TXML['Rmonth']>6) | ((TXML['Rmonth']==6) & (TXML['Rday']>2)))].copy()
#             TXML[['Rmonth', 'Rday']] = TXML[['Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
#             TXML = TXML[((TXML['Rmonth']<9) | ((TXML['Rmonth']==9) & (TXML['Rday']<3)))].copy()
#             TXML['Tyear'] = pd.to_numeric(TXML['Tyear'], errors='coerce')
#             TXML = TXML[TXML['Tyear']==994].copy()
            
#             #TX CURR
#             df['Ryear'] = pd.to_numeric(df['Ryear'], errors='coerce')
#             a = df[df['Ryear']==2025].copy()
#             a['Tyear'] = pd.to_numeric(a['Tyear'], errors='coerce')
#             a = a[a['Tyear']==994].copy()
            
#             b = df[df['Ryear']==2024].copy()
#             b[['Rmonth', 'Rday']] = b[['Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
#             b = b[((b['Rmonth']>9) | ((b['Rmonth']==9) & (b['Rday']>2)))].copy()
#             b['Tyear'] = pd.to_numeric(b['Tyear'], errors='coerce')
#             b = b[b['Tyear']==994].copy()
#             TXCURR = pd.concat([a,b])
            
#             #TX NEW
#             df[['Ayear', 'Amonth']] = df[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
#             TXNEW = df[((df['Ayear']==2024) & (df['Amonth'].isin([7,8,9])))].copy()
#             df[['Tiyear', 'Timonth']] = df[['Tiyear', 'Timonth']].apply(pd.to_numeric, errors='coerce')
#             TI = df[((df['Tiyear']==2024) & (df['Timonth'].isin([7,8,9])))].copy()
            

#             #TRANSFER OUTS
#             df[['Tyear', 'Tmonth']] = df[['Tyear', 'Tmonth']].apply(pd.to_numeric, errors='coerce')
#             TO = df[df['Tyear']!=994].copy()
#             TO[['Ryear', 'Rmonth', 'Rday']] = TO[['Ryear', 'Rmonth','Rday']].apply(pd.to_numeric, errors='coerce')
            
#             TOa = TO[((TO['Ryear']==2024) & (TO['Rmonth']<10))].copy()
#             TOa[['Rmonth', 'Rday']] = TOa[['Rmonth','Rday']].apply(pd.to_numeric, errors='coerce')
#             #st.write(TOa)
#             TOa = TOa[((TOa['Rmonth'] >6) | ((TOa['Rmonth'] ==6) & (TOa['Rday'] >2)))].copy()
#             #TOa[['Tmonth', 'Tyear']] = TOa[['Tmonth','Tyear']].apply(pd.to_numeric, errors='coerce')
#             #TOa = TOa[((TOa['Tyear']==2024) & (TOa['Tmonth'].isin([4,5,6])))].copy()

#             TO[['Ryear', 'Rmonth', 'Rday']] = TO[['Ryear', 'Rmonth','Rday']].apply(pd.to_numeric, errors='coerce')
#             FALSE = TO[((TO['Ryear']>2024) | ((TO['Ryear']==2024) & (TO['Rmonth']>9)))].copy()

#             TXCUR = pd.concat([TXCURR,FALSE])

#             #VL COV
#             TXCUR['Ayear'] = pd.to_numeric(TXCUR['Ayear'], errors='coerce')
#             c = TXCUR[ TXCUR['Ayear']==2024].copy()
#             d = TXCUR[ TXCUR['Ayear']<2024].copy()
#             d[['Vyear', 'Vmonth']] = d[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
#             e = d[((d['Vyear'] ==2024) | ((d['Vyear'] ==2023) & (d['Vmonth'] >9)))].copy()
#             f = d[((d['Vyear'] < 2023) | ((d['Vyear'] ==2023) & (d['Vmonth'] <10)))].copy()
#             WVL = pd.concat([c,e])
            
#             NOVL = f.copy()
#             cphl = r'AVLS.csv'
#             cp = pd.read_csv(cphl)

#             POTENTIAL = potential
#             newad = TXNEW.shape[0]
#             out = TOa.shape[0]
#             inn = TI.shape[0]
#             curr = TXCUR.shape[0]
#             false = FALSE.shape[0]
#             lost = TXML.shape[0]
#             vl = WVL.shape[0]
#             #st.write(vl)
#             #st.write(curr)
#             #st.stop()
#             perc = round((vl/curr)*100)
#             exp = round(curr*0.95)
#             novl = NOVL.shape[0]
#             current_time = time.localtime()
#             week = time.strftime("%V", current_time)
#             week = int(week) +13

#             districts = list(dfx['DISTRICT'].unique())
#             district = st.radio(label='**Choose a district**', options=districts,index=None, horizontal=True)
#             if district:
#                 facilities = dfx[dfx['DISTRICT']==district]
#                 facilities = facilities['FACILITY']
#                 facility = st.selectbox(label='**Choose a facility**', options=facilities,index=None)
#                 if facility:
#                     preva = dfx[dfx['FACILITY'] == facility]
#                     prev = int(preva.iloc[0,3])
#                     name =str(preva.iloc[0,4])
#                     UK = potential- prev - inn - newad
#                     dd = dead.shape[0]
#                     if UK <0:
#                         st.warning('THIS EXTRACT HAS LESS CLIENTS THAN EVER ENROLLED AT THE FACILITY')
#                         st.stop()
#                     else:
#                         pass
#                     ba = prev - curr
#                     if ba > 0:
#                         bal = ba
#                     elif ba == 0:
#                         bal = 'EVEN'
#                     elif ba < 0:
#                         bal = 'EXCEEDED'
#                     grow = curr-prev
#                     if grow ==0:
#                         st.success(f'WEBALE {name},😐 this TXCURR has broken even (Q3 CURR is equal to Q4 CURR), but you need to add more clients to grow it even further 👏👏👏')
#                         if perc > 94:
#                             st.success(f'Even the VL COVERAGE is good, at {perc}%  👏👏👏')
#                         else:
#                             st.warning(f'**However the VL COVERAGE is poor, at {perc}%** 🥲')

#                     elif grow>0:
#                         st.success(f'WEBALE {name},😐 you have grown this TXCURR by {grow}, but you need to audit the TIs and TXNEWs, and watch out for RTT 👏👏👏')
#                         if perc > 94:
#                             st.success(f'Even the VL COVERAGE is good, at {perc}%  👏👏👏')
#                             st.balloons()
#                             time.sleep(2)
#                             st.balloons()
#                             time.sleep(2)
#                             st.balloons()
#                             time.sleep(2)
#                             st.balloons()
#                             time.sleep(2)
                            
#                         else:
#                             st.warning(f'**However the VL COVERAGE is poor, at {perc}%** 🥲')
#                             st.balloons()
#                             time.sleep(2)
#                             st.balloons()                       
#                     else:
#                         st.warning(f'**BANANGE {name}, 😢 you have dropped this TXCURR by {grow}, you need to audit the TXMLs and TOs, and watch out for the dead** 😢😢😢')
#                         if perc > 94:
#                             st.success(f'BUT the VL COVERAGE is good, at {perc}% 👏')
#                         else:
#                             st.warning(f'**EVEN the VL COVERAGE is poor, at {perc}%** 😢😢😢')
#                     cp = cp[cp['facility']==facility].copy()
#                     cp = cp.rename(columns ={'ART-NUMERIC': 'ART'})
#                     cp['ART'] = pd.to_numeric(cp['ART'], errors='coerce')
#                     NOV = NOVL[['ART', 'RETURN DATE', 'VL DATE']].copy()
#                     NOV['ART'] = pd.to_numeric(NOV['ART'], errors='coerce')

#                     AT = pd.merge(cp, NOV, on='ART',how='inner')
#                     NOVL['ART'] = pd.to_numeric(NOVL['ART'], errors='coerce')
                    
#                     AT['ART'] = pd.to_numeric(AT['ART'], errors='coerce')
#                     TRUE = NOVL[~NOVL['ART'].isin(AT['ART'])].copy()
                    
#                     oneyear[['Ayear', 'Amonth']] = oneyear[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
#                     new = oneyear[((oneyear['Ayear']==2023) & (oneyear['Amonth'].isin([7,8,9])))].copy()
    
#                     new[['Tiyear']] = new[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
#                     tin = new[new['Tiyear']!=994].copy()
#                     #one =new.shape[0]
#                     tew = tin.shape[0]
                    
#                     newtotal = new.shape[0]
                    
#                     new['Dyear'] = pd.to_numeric(new['Dyear'], errors='coerce')
#                     newdead = new[new['Dyear']!=994].copy()

#                     deadnew = newdead.shape[0]
#                     new = new[new['Dyear']==994].copy()

#                     new['Tyear'] = pd.to_numeric(new['Tyear'], errors='coerce')
                    
#                     newto = new[new['Tyear']!=994].copy()
#                     outnew = newto.shape[0]
                    
#                     new = new[new['Tyear']==994].copy()
#                     new['ART'] = pd.to_numeric(new['ART'], errors = 'coerce')
#                     TXCUR['ART'] = pd.to_numeric(TXCUR['ART'], errors = 'coerce')
                    
#                     active = new[new['ART'].isin(TXCUR['ART'])].copy()
#                     lostn = new[~new['ART'].isin(TXCUR['ART'])].copy()
                    

#                     newactive = active.shape[0]
#                     newlost = lostn.shape[0]
#                     #st.write(newlost)
                           
#                     #ret = newtotal - newlost
#                     if newtotal == 0:
#                         rete = 0
#                     elif newactive == 0:
#                         rete = 0
#                     else:
#                         rete = round((newactive/newtotal)*100)
#                         rete = f"{rete} %"
                    
#                     data = pd.DataFrame([{
#                                 'DISTRICT': district,
#                                 'FACILITY' : facility,
#                                 'Q3 CURR':prev,
#                                 'UNKNOWN GAIN': UK,
#                                 'DEAD': dd,
#                                 'POTENTIAL': potential,
#                                 'Q4 CURR': curr,
#                                 'TXML' : lost,
#                                  'BALANCE': bal,
#                                 'TX NEW' : newad,
#                                 'TO' : out,
#                                 'FALSE TO': false,
#                                 'TI': inn,
#                                 'HAS VL' : vl,
#                                 'VL COV (%)': perc,
#                                 'EXPECTED': exp,
#                                 'NO VL' : novl,
#                                 'WEEK': week,
#                                  'ORIGINAL COHORT': newtotal,
#                                  'ONE YEAR TI': tew,
#                                  'ONE YEAR LOST': newlost,
#                                   'ONE YEAR TO': outnew,
#                                  'ONE YEAR DEAD': deadnew,
#                                  'ONE YEAR ACTIVE': newactive,
#                                  'ONE YR RETENTION': rete
#                                  }])
#                     #data = data.set_index('DISTRICT')
                    
#                     #SUBMISSION
#                     # conn = st.connection('gsheets', type=GSheetsConnection)
#                     # exist = conn.read(worksheet ='TXML', usecols = list(range(15)), ttl=5)
#                     # existing = exist.dropna(how='all')
#                     col1,col2,col3 = st.columns([1,2,1])
#                     with col3:
#                         submit = st.button('Submit') 
                      
#                     if submit:
#                         try:
#                             pass
#                             # #conn = st.connection('gsheets', type=GSheetsConnection)
#                             # exist = conn.read(worksheet ='TXML', usecols = list(range(25)), ttl=5)
#                             # existing = exist.dropna(how='all')
#                             # updated = pd.concat([existing, data], ignore_index =True)
#                             # conn.update(worksheet = 'TXML', data = updated)
#                             # st.success('Your data above has been submitted')
#                         except:
#                             st.write("Couldn't submit, poor network")
#                     if submit:
#                         st.write('**TX CURR AS OF 3rd SEPT**')
#                         pass
#                         st.session_state.submited = True
#                     else:
#                         st.write('')
#                     if st.session_state.submited:
#                         st.dataframe(data)
#                         st.write(f"<h6>DOWNLOAD LINELISTS FROM HERE</h6>", unsafe_allow_html=True)
#                         cola, colb, colc = st.columns(3)
#                         with cola:
#                              dat = TXCUR.copy()
                             
#                              dat = dat[['ART NO', 'ART START DATE', 'RETURN DATE', 'VL DATE']]
#                              csv_data = dat.to_csv(index=False)
#                              st.download_button(
#                                          label=" DOWNLOAD TXML",
#                                          data=csv_data,
#                                          file_name=f"{facility} TXML.csv",
#                                          mime="text/csv")
#                         with colb:
#                              dat = NOVL.copy()
#                              dat = dat[['ART NO', 'ART START DATE', 'RETURN DATE', 'VL DATE']]
                             
#                              csv_data = dat.to_csv(index=False)
#                              st.download_button(
#                                              label=" DOWNLOAD WITH NO VL",
#                                              data=csv_data,
#                                              file_name=f" {facility} NO VL.csv",
#                                              mime="text/csv")
#                         with colc:
#                              dat = TOa.copy()
#                              dat = dat[['ART NO', 'ART START DATE', 'RETURN DATE', 'VL DATE', 'T OUT DATE']]
#                              #dat = AT.copy()
#                              csv_data = dat.to_csv(index=False)
#                              st.download_button(
#                                          label=" DOWNLOAD TRANSFER OUTS",
#                                          data=csv_data,
#                                          file_name=f" {facility} T OUTS.csv",
#                                          mime="text/csv")
    
#     #########################################################################################################################################################
#                         AT= AT[['ART', 'RETURN DATE', 'VL DATE','art_number','date_collected','result_numeric']].copy()
#                         a = AT.shape[0]
#                         if a==0:   
#                             st.write('**I DO NOT SEE VL RESULTS AT CPHL MISSING IN THIS EMR EXTRACT FOR NOW. WAIT FOR FUTURE UPDATES**')
#                         elif a == 1:
#                             st.success(f'I see only **{a}** result at CPHL that is not yet entered into EMR')
#                         else:
#                             st.success(f'I see over **{a}** results at CPHL that are not yet entered into EMR')
                                
            
#                         cola, colb = st.columns([2,1])
#                         with cola:
#                             if a>0:  
#                                 named = facility
#                                 #if st.button('DOWNLOAD FILE'):
#                                 wb = Workbook()
#                                 ws = wb.active
                     
#                         # Convert DataFrame to Excel
#                                 for r_idx, row in enumerate(AT.iterrows(), start=1):
#                                        for c_idx, value in enumerate(row[1], start=1):
#                                             ws.cell(row=r_idx, column=c_idx, value=value)
#                                 ws.insert_rows(0,2)
                    
#                                 blue = PatternFill(fill_type = 'solid', start_color = 'F6F8F7')
#                                     # ws.column_dimensions['H'].width = 14
                    
#                                 for num in range (1, ws.max_row+1):
#                                      for letter in ['D','E', 'F']:
#                                           ws[f'{letter}{num}'].font = Font(b= True, i = True)
#                                           ws[f'{letter}{num}'].font = Font(b= True, i = True)
#                                           ws[f'{letter}{num}'].fill = blue
#                                           ws[f'{letter}{num}'].border = Border(top = Side(style = 'thin', color ='000000'),
#                                                                                 right = Side(style = 'thin', color ='000000'),
#                                                                                 left = Side(style = 'thin', color ='000000'),
#                                                                                 bottom = Side(style = 'thin', color ='000000'))
#                                 ws['B1'] ='EMR DETAILS'
#                                 ws['F1'] = 'CPHL DETAILS'
#                                 ws['A2'] = 'ART-NO'
#                                 ws['B2'] = 'RETUR VISIT DATE'
#                                 ws['C2'] = 'EMR VL DATE' 
#                                 ws['D2'] = 'ART NO'
#                                 ws['E2'] = 'CPHL DATE'
#                                 ws['F2']  = 'CPHL RESULTS'
                    
                    
#                                 letters = ['B', 'C', 'D','E','F']
#                                 for letter in letters:
#                                       ws.column_dimensions[letter].width =15
                    
#                                 ran = random.random()
#                                 rand = round(ran,2)
#                                 file_path = os.path.join(os.path.expanduser('~'), 'Downloads', f'{named}_missing_results {rand}.xlsx')
#                                 directory = os.path.dirname(file_path)
#                                 Path(directory).mkdir(parents=True, exist_ok=True)
                    
#                                       # Save the workbook
#                                 wb.save(file_path)
#                          # Serve the file for download
#                                 with open(file_path, 'rb') as f:
#                                   file_contents = f.read()           
#                                 st.download_button(label=f'DONLOAD MISSING RESULTS FOR {named} ', data=file_contents,file_name=f'{named}_missing_results {rand}.xlsx', 
#                                                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#                             else:
#                                 pass
#                         with colb:
#                              if dd >0:             
#                                  st.write('**CONFIRM IF THESE ARE DEAD**')
#                                  dat = dead.copy()
#                                  #st.write(dead.columns)
#                                  dat = dat[['A', 'RD', 'DD']]
#                                  dat = AT.copy()
#                                  csv_data = dat.to_csv(index=False)
#                                  st.download_button(
#                                              label=" DOWNLOAD DEAD LINELIST",
#                                              data=csv_data,
#                                              file_name=f" {facility} DEAD.csv",
#                                              mime="text/csv")
#                              else:
#                                  pass
                                
#     #########################################################################################################
#                         #st.write(lostn)
#                         st.write('**ONE YEAR COHORT RETENTION**')
#                         one = data[['ORIGINAL COHORT','ONE YEAR LOST','ONE YEAR DEAD', 'ONE YEAR TO' ,'ONE YEAR ACTIVE',  'ONE YR RETENTION']].copy()
                                 
#                         one = one.rename(columns ={'ONE YEAR LOST': 'LOST','ONE YEAR DEAD': 'DEAD','ONE YEAR TO':'TO',  'ONE YEAR ACTIVE': 'ACTIVE'})
#                         one = one.set_index('ORIGINAL COHORT')
#                         st.write(one)
#                         cola,colb = st.columns(2)
#                         if newlost==0:
#                             st.write('**NO IIT AMONGST LOST CLIENTS, SO NO LINE LIST TO DOWNLOAD**')
#                             pass
#                         else:         
#                             lostn = lostn.rename(columns ={'A':'ART NO', 'AS':'ART START DATE', 'RD':'RETURN DATE', 'VD':'VL DATE'})
#                             with cola:
#                                  dat = lostn.copy()
#                                  #dat = TXCUR.copy()
#                                  dat = dat[['ART NO', 'ART START DATE','RETURN DATE', 'TI']]
#                                  csv_data = dat.to_csv(index=False)
#                                  st.download_button(
#                                              label="DOWNLOAD_IIT_FOR_1_YR_COHORT",
#                                              data=csv_data,
#                                              file_name=f"{facility}_1_YR_IIT.csv",
#                                              mime="text/csv")
    
#                         if   outnew ==0:
#                             pass
#                         else:
#                              with colb:
#                                  dat =  newto.copy()
#                                  dat = dat[['A', 'AS','RD', 'TO','TI']]
#                                  csv_data = dat.to_csv(index=False)
#                                  st.download_button(
#                                              label="DOWNLOAD_TOs_FOR_1_YR_COHORT",
#                                              data=csv_data,
#                                              file_name=f"{facility}_1_YR_TOs.csv",
#                                              mime="text/csv")
#                     else:
#                         st.write('**FIRST SUBMIT TO SEE THE LINE-LISTS**')
#                         st.stop()
                        

                                 
                                
                    
